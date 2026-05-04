"""Parse Office files (.docx, .xlsx) into structured nodes."""

from __future__ import annotations

from pathlib import Path

from gathon.schema import Confidence, FileType, UnifiedEdge, UnifiedNode


def parse_office(path: Path) -> tuple[list[UnifiedNode], list[UnifiedEdge]]:
    """Route to docx or xlsx parser."""
    if path.suffix.lower() == ".docx":
        return _parse_docx(path)
    elif path.suffix.lower() == ".xlsx":
        return _parse_xlsx(path)
    return [], []


def _parse_docx(path: Path) -> tuple[list[UnifiedNode], list[UnifiedEdge]]:
    """Parse .docx → heading hierarchy."""
    try:
        import docx
    except ImportError:
        return [], []

    nodes = []
    edges = []

    root = UnifiedNode(
        kind="Document",
        name=path.stem,
        qualified_name=f"{path}::root",
        file_path=str(path),
        label=path.stem,
        file_type=FileType.DOCUMENT,
        confidence=Confidence.EXTRACTED,
        confidence_score=1.0,
        pipeline="gathon_office",
    )
    nodes.append(root)

    try:
        doc = docx.Document(path)
    except Exception:
        return [root], []

    stack = [(0, root)]

    for para in doc.paragraphs[:100]:
        style = para.style.name if para.style else ""
        if not style.startswith("Heading"):
            continue

        level = int(style.split()[-1]) if style[-1].isdigit() else 1
        text = para.text.strip()
        if not text:
            continue

        section = UnifiedNode(
            kind="Section",
            name=text[:50],
            qualified_name=f"{path}::h{len(nodes)}",
            file_path=str(path),
            label=text,
            file_type=FileType.DOCUMENT,
            confidence=Confidence.EXTRACTED,
            confidence_score=1.0,
        )
        nodes.append(section)

        while len(stack) > 1 and stack[-1][0] >= level:
            stack.pop()

        parent = stack[-1][1]
        edges.append(UnifiedEdge(
            kind="CONTAINS",
            source_qualified=parent.qualified_name,
            target_qualified=section.qualified_name,
            file_path=str(path),
            relation="contains",
            confidence=1.0,
        ))

        stack.append((level, section))

    return nodes, edges


def _parse_xlsx(path: Path) -> tuple[list[UnifiedNode], list[UnifiedEdge]]:
    """Parse .xlsx → sheet nodes with column headers."""
    try:
        import openpyxl
    except ImportError:
        return [], []

    nodes = []
    edges = []

    root = UnifiedNode(
        kind="Document",
        name=path.stem,
        qualified_name=f"{path}::root",
        file_path=str(path),
        label=path.stem,
        file_type=FileType.CONFIG,
        confidence=Confidence.EXTRACTED,
        confidence_score=1.0,
        pipeline="gathon_office",
    )
    nodes.append(root)

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return [root], []

    for sheet in wb.sheetnames[:10]:
        ws = wb[sheet]

        sheet_node = UnifiedNode(
            kind="Section",
            name=sheet,
            qualified_name=f"{path}::sheet_{sheet}",
            file_path=str(path),
            label=f"Sheet: {sheet}",
            file_type=FileType.CONFIG,
            confidence=Confidence.EXTRACTED,
            confidence_score=1.0,
        )
        nodes.append(sheet_node)

        edges.append(UnifiedEdge(
            kind="CONTAINS",
            source_qualified=root.qualified_name,
            target_qualified=sheet_node.qualified_name,
            file_path=str(path),
            relation="contains",
            confidence=1.0,
        ))

        for i, cell in enumerate(ws[1][:20]):
            if not cell.value:
                continue
            col_node = UnifiedNode(
                kind="ConfigKey",
                name=str(cell.value),
                qualified_name=f"{path}::col_{sheet}_{i}",
                file_path=str(path),
                label=str(cell.value),
                file_type=FileType.CONFIG,
                confidence=Confidence.EXTRACTED,
                confidence_score=1.0,
            )
            nodes.append(col_node)

            edges.append(UnifiedEdge(
                kind="CONTAINS",
                source_qualified=sheet_node.qualified_name,
                target_qualified=col_node.qualified_name,
                file_path=str(path),
                relation="contains",
                confidence=1.0,
            ))

    return nodes, edges
